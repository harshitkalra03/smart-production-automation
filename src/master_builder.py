from pathlib import Path
import json
from openpyxl import load_workbook


class MasterBuilder:

    def __init__(self):

        self.project_root = Path(__file__).resolve().parent.parent

        self.data_dir = self.project_root / "data"

        self.brand_file = self.data_dir / "Brand Models.xlsx"
        self.tank_file = self.data_dir / "Tank_Types.xlsx"
        self.shape_file = self.data_dir / "Model_Shape.xlsx"
        self.power_file = self.data_dir / "Model_Power.xlsx"

        self.master_file = self.data_dir / "master_products.json"

        self.brand_headers = {
            "ORIENT","KENSTAR","SURYA","POLYCAB","CROMA","HINDWARE",
            "RR","GOLDMEDAL","LAZER","KROMO","SUNFLAME","VETO",
            "JOHNSON","GM ELEKTRA","BPL","VENUS","BLACK+DECKER",
            "LIFELONG","GLEN","MCCOY","ACCURA","ALLOMAX","INALSA",
            "SMART","SAKASH","SUJATA","SAMEER", "POLAR"
        }

        self.tank_headers = {
            "SS TANK",
            "POLYMER",
            "GLASSLINE"
        }

        self.shape_headers = {
            "HORIZONTAL",
            "INSTANT",
            "SQUARE",
            "VERTICAL"
        }

        self.power_headers = {
            "2 KW",
            "3 KW",
            "4.5 KW"
        }

    def normalize_model(self, model):

        model = str(model).strip()

        special_cases = {
            "INSTANT": "INSTANT_IWH",
            "3L,4.5K KW": "3L_4P5KW"
        }

        if model.upper() in special_cases:
            return special_cases[model.upper()]

        return model.title()

    def read_column(self, file_path, column_index):

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        values = []

        for row in ws.iter_rows():

            cell = row[column_index]

            if cell.value is None:
                continue

            value = str(cell.value).strip()

            if value:
                values.append(value)

        return values

    def build_brand_map(self):

        data = self.read_column(self.brand_file, 0)

        current_brand = None
        mapping = {}

        for item in data:

            if item.upper() in self.brand_headers:
                current_brand = item.title()
                continue

            if current_brand:
                mapping[self.normalize_model(item)] = current_brand

        return mapping

    def build_tank_map(self):

        data = self.read_column(self.tank_file, 0)

        current_tank = None
        mapping = {}

        for item in data:

            if item.upper() in self.tank_headers:
                current_tank = item.title()
                continue

            if current_tank:
                mapping[self.normalize_model(item)] = current_tank

        return mapping

    def build_shape_map(self):

        data = self.read_column(self.shape_file, 0)

        current_shape = None
        mapping = {}

        for item in data:

            if item.upper() in self.shape_headers:
                current_shape = item.title()
                continue

            if current_shape:
                mapping[self.normalize_model(item)] = current_shape

        return mapping

    def build_power_map(self):

        data = self.read_column(self.power_file, 0)

        current_power = None
        mapping = {}

        for item in data:

            normalized = item.upper().replace(" ", "")

            if normalized in {"2KW", "3KW", "4.5KW"}:
                current_power = item
                continue


            if current_power:
                mapping[self.normalize_model(item)] = current_power

        return mapping

    def build_master(self):

        brand_map = self.build_brand_map()
        tank_map = self.build_tank_map()
        shape_map = self.build_shape_map()
        power_map = self.build_power_map()

        all_models = (
            set(brand_map.keys())
            | set(tank_map.keys())
            | set(shape_map.keys())
            | set(power_map.keys())
        )

        master = {}

        missing_brand = 0
        missing_tank = 0
        missing_shape = 0
        missing_power = 0

        missing_brand_models = []
        missing_tank_models = []
        missing_shape_models = []
        missing_power_models = []

        for model in sorted(all_models):

            brand = brand_map.get(model, "")
            tank = tank_map.get(model, "")
            shape = shape_map.get(model, "")
            power = power_map.get(model, "")

            if not brand:
                missing_brand += 1
                missing_brand_models.append(model)

            if not tank:
                missing_tank += 1
                missing_tank_models.append(model)

            if not shape:
                missing_shape += 1
                missing_shape_models.append(model)

            if not power:
                missing_power += 1
                missing_power_models.append(model)

            master[model] = {
                "brand": brand,
                "tank_type": tank,
                "shape": shape,
                "power": power
            }

        # Save Master JSON
        with open(self.master_file, "w", encoding="utf-8") as f:
            json.dump(
                master,
                f,
                indent=4,
                ensure_ascii=False
            )

        # Validation Report
        validation_file = self.data_dir / "validation_report.txt"

        with open(validation_file, "w", encoding="utf-8") as f:

            f.write("MASTER DATA VALIDATION REPORT\n")
            f.write("=" * 60 + "\n\n")

            f.write(f"Total Models : {len(master)}\n\n")

            f.write(f"Missing Brand : {missing_brand}\n")
            for item in missing_brand_models:
                f.write(f"  - {item}\n")

            f.write("\n")

            f.write(f"Missing Tank : {missing_tank}\n")
            for item in missing_tank_models:
                f.write(f"  - {item}\n")

            f.write("\n")

            f.write(f"Missing Shape : {missing_shape}\n")
            for item in missing_shape_models:
                f.write(f"  - {item}\n")

            f.write("\n")

            f.write(f"Missing Power : {missing_power}\n")
            for item in missing_power_models:
                f.write(f"  - {item}\n")

        print("\nMASTER BUILD COMPLETE")
        print("-" * 50)
        print(f"Total Models : {len(master)}")
        print(f"Missing Brand : {missing_brand}")
        print(f"Missing Tank  : {missing_tank}")
        print(f"Missing Shape : {missing_shape}")
        print(f"Missing Power : {missing_power}")

        print("\nValidation Report Generated")
        print(f"Location : {validation_file}")
    def run(self):
        self.build_master()


if __name__ == "__main__":
    MasterBuilder().run()
    